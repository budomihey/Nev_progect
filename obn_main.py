import numpy as np

#'Переменные для работы'
#'Qinj1-2,Pb1-2,Pk-2, Pz1-2,D1-2'

import openpyxl
#а вот теперь подгружаем наконец эксельку
wb = openpyxl.load_workbook(filename= 'C:\лишнеее\kagbekursach_kt-main\BMA.xlsm')
sheet = wb['571 вгп']

val = sheet['A1'].value
print(val)
b = sheet.cell(row=5, column=10).value
print(b)
j=0
Qing=np.zeros(31)
Pb=np.zeros(31)
D=np.zeros(31)
Pk=np.zeros(31)
Kl=np.zeros(31)
Pz=np.zeros(31)
kq=0.1


while j<31:
    Qing[j]=sheet.cell(row=5, column=j+10).value
    Pb[j]=sheet.cell(row=7, column=j+10).value
    D[j]=(sheet.cell(row=10, column=j+10).value)
    Pk[j]=(sheet.cell(row=12, column=j+10).value)
    Kl[j]=(sheet.cell(row=15, column=j+10).value)
    Pz[j]=(sheet.cell(row=8, column=j+10).value)
    j = j+1
j=0
while j<30:
    if (Qing[j]==0 and Pk[j] != 0) or (Qing[j] != 0 and Pk[j] == 0):
        Qing[j]=Qing[j-1]
        Pk[j]=Pk[j-1]
    elif Qing[j]==0 and Pk[j]==0:
        j=30
    if ((Qing[j] - Qing[j-1]) >(kq*Qing[j-1])) and ((Pk[j]-Pk[j-1])>(kq*Pk[j-1])):
        if Kl[j]-Kl[j-1]>kq*Kl[j-1]:
            print('у нас АвтоГРП или авария, проверить Рзатр')
            if Pz[j]-Pz[j-1]>kq*Pz[j-1]:
                print('Авария пакера, провести ревизию пакера!')
            elif Pz[j]-Pz[j-1]<=kq*Pz[j-1]:
                print('АвтоГРП, сменить режим работы скважины')

        elif abs(Kl[j]-Kl[j-1])<kq*Kl[j-1]:
            print('Норма, смена технологического режима')
        elif Kl[j]-Kl[j-1]<(-1)*kq*Kl[j-1]:
            print('Кольматация на смене режима, для уточнения проверить Рбуф')
            if Qing[j]/Pb[j]-Qing[j-1]/Pb[j-1]<(-1)*kq*Qing[j-1]/Pb[j-1]:
                print('Кольматация, требуется обработка ПЗП, пнуть КРСников на это')
            elif Qing[j]/Pb[j]-Qing[j-1]/Pb[j-1]>=(-1)*kq*Qing[j-1]/Pb[j-1]:
               print('ошибка систем телеметрии, провести ревизию')
    if Qing[j] - Qing[j - 1] > kq * Qing[j - 1] and abs(Pk[j] - Pk[j - 1]) < kq * Pk[j - 1]:
        print('Проверить Рбуф, без него уточнение не получится')
        if Qing[j] - Qing[j+1] > Qing[j]*(Pk[j]/Pb[j]-1):
            print('Либо автоГРП, либо авария пакера, проверить Рзатр')
            if Pz[j]-Pz[j-1]>kq*Pz[j-1]:
                print('Авария пакера,  провести ревизию ПО')
            else:
                print('АвтоГРП, сменить режим работы скважины ')
        elif Qing[j] - Qing[j+1] <= Qing[j]*(Pk[j]/Pb[j]-1):
            print('вполне возможна кольматация, проверить Рбуф и Dшт')
            if D[j]!=D[j-1]:
                print('Штуцирование скважины, провести ревизию штуцера')
            else:
                if Pb[j]-Pb[j+1]>kq*Pb[j-1]:
                    print('Штуцер размыт, заменить')
                else:
                    print('ошибка систем телеметрии, провести ревизию')
    if Qing[j] - Qing[j - 1] > kq * Qing[j - 1] and (Pk[j] - Pk[j - 1]) < (-1)*kq * Pk[j - 1]:
        print('АвтоГРП или авария пакера, уточнить Рзатр')
        if Pz[j] - Pz[j - 1] > kq * Pz[j - 1]:
            print('Авария пакера,  провести ревизию ПО')
        else:
            print('АвтоГРП, сменить режим работы скважины')
    if abs(Qing[j] - Qing[j - 1]) < kq * Qing[j - 1] and (Pk[j] - Pk[j - 1]) > kq * Pk[j - 1]:
        print('Вероятнее всего кольматация ПЗП, проверить Рбуф')
        if Qing[j]/Pb[j]-Qing[j-1]/Pb[j+1]>kq*Qing[j-1]/Pb[j-1]:
            print('Кольматация, провести обработку ПЗП')
        else:
            print('неясные проблемы, возможно телеметрия в отпуске')
    if abs(Qing[j] - Qing[j - 1]) < kq * Qing[j - 1] and (Pk[j] - Pk[j - 1]) <(-1)* kq * Pk[j - 1]:
        print('Возможны АвтоГРП или авария, уточнить Рзатр')
        if Pz[j] - Pz[j - 1] > kq * Pz[j - 1]:
            print('Авария пакера, КРС к бою, провести ревизию')
        else:
            print('АвтоГРП, кореектируем режим работы скважины')
    if (Qing[j] - Qing[j - 1]) <(-1)* kq * Qing[j - 1] and (Pk[j] - Pk[j - 1]) < (-1) * kq * Pk[j - 1]:
        if Kl[j]-Kl[j-1]<(-1)*kq*Kl[j-1]:
            print('Смена режима работы скважины, норма')
        else:
            print('Либо кольматация, либо штуцирование, проверить Dшт и Рбуф')
            if D[j]<D[j-1]and Pb[j]-Pb[j-1]<(-1)*kq*Pb[j-1]:
                print('Штуцирование, выставить штуцер (на мороз)')
            elif (D[j]>D[j-1]and Pb[j]-Pb[j-1]<(-1)*kq*Pb[j-1]) or (D[j]<D[j-1]and Pb[j]-Pb[j-1]>kq*Pb[j-1]):
                print('Ошибка при измерении или авария штуцера, провести ревизию штуцера и телеметрии')
            elif D[j]!=D[j-1] and Pb[j]-Pb[j-1]<(-1)*kq*Pb[j-1]:
                print('Кольматация, провести ПЗП')
    j=j+1








  #Пока мат и шутки почти убраны